# Copyright (c) 2026, Falendra Bandhe and contributors
# For license information, please see license.txt

"""Phase 5: deterministic import of data/source/*.xlsx|csv into the live site.

Run via:
    bench --site <site> execute california_burrito.utils.import_data.run \\
        --kwargs "{'source_dir': '/workspace-project/data/source'}"

Dependency order per docs/DocType_Spec.md's header:
    Zonal Office -> Outlet -> Asset Type -> Technician -> Asset -> PM Program ->
    Ticket Taxonomy -> Spare Part -> (seed initial PM Schedule rows)

Idempotent by construction: every doctype here is get-or-create against its natural
key (outlet_code, employee_no, asset_type_name, program_key, taxonomy_key, part_code),
and schedule seeding goes through ensure_schedule, which is idempotent on
generation_key. Re-running this against an already-imported site creates nothing new
except logging "already exists, skipped" — it does not error or duplicate.

Every persistent record this project already had before this import (the Phase 1-4
fixture/demo data: CB Outlet BLR001 and BLR134, CB Asset AST-00001..00005, CB
Technician DEMO-TECH-01, CB Asset Type Air Conditioner/Walk-in Chiller/Freezer/Fryer,
CB PM Program "Air Conditioner / Clean filter / Monthly") is left untouched — every
insert here is checked against frappe.db.exists() on the real natural key first, so
nothing here can collide with or duplicate it.
"""

import csv
import os

import frappe
from frappe.utils import today

from california_burrito.utils.normalization import (
	ASSET_TYPE_ALIASES,
	canonicalize_asset_type,
	match_reports_to,
	normalize_frequency,
	normalize_whitespace,
	parse_spare_part,
)
from california_burrito.utils.schedule import ensure_schedule, find_applicable_targets

# CB Outlet.city only has the 6 store-serving cities — no outlet is a Corporate
# office. CB Zonal Office.city additionally has COR (added in Phase 5: one real
# technician's Home is "COR", not a city — see PROGRESS.md and docs/DocType_Spec.md's
# amendment note on CB Zonal Office).
OUTLET_CITIES = ("NCR", "BLR", "HYD", "CHN", "PUN", "MUM")
ZONAL_OFFICE_CITIES = OUTLET_CITIES + ("COR",)

HOME_TO_CITY = {
	"Zonal Office - Delhi/NCR": "NCR",
	"Zonal Office - Bengaluru": "BLR",
	"Zonal Office - Hyderabad": "HYD",
	"Zonal Office - Chennai": "CHN",
	"Zonal Office - Pune": "PUN",
	"Zonal Office - Mumbai": "MUM",
	"COR": "COR",
}


class ImportSummary:
	def __init__(self):
		self.counts = {}
		self.warnings = []

	def count(self, label, n=1):
		self.counts[label] = self.counts.get(label, 0) + n

	def warn(self, message):
		self.warnings.append(message)

	def print_report(self):
		print("\n" + "=" * 78)
		print("PHASE 5 IMPORT SUMMARY")
		print("=" * 78)
		for label, n in self.counts.items():
			print(f"  {label}: {n}")
		print(f"\nWarnings ({len(self.warnings)}):")
		for w in self.warnings:
			print(f"  - {w}")
		print("=" * 78)


def _read_xlsx_rows(path, sheet_name=None):
	import openpyxl

	wb = openpyxl.load_workbook(path, data_only=True)
	ws = wb[sheet_name] if sheet_name else wb[wb.sheetnames[0]]
	header = [normalize_whitespace(c.value) for c in ws[1]]
	return [dict(zip(header, row)) for row in ws.iter_rows(min_row=2, values_only=True)]


def import_zonal_offices(summary):
	"""One Zonal Office per real city option, plus Corporate Office (COR — added in
	Phase 5). Reuses 'BLR Zonal Office' from the Phase 1 fixture rather than creating
	a duplicate for BLR."""
	zonal_office_by_city = {}
	for city in ZONAL_OFFICE_CITIES:
		name = "Corporate Office" if city == "COR" else f"{city} Zonal Office"
		if not frappe.db.exists("CB Zonal Office", name):
			frappe.get_doc({"doctype": "CB Zonal Office", "zonal_office_name": name, "city": city}).insert()
			summary.count("CB Zonal Office created")
		zonal_office_by_city[city] = name
	return zonal_office_by_city


def import_outlets(summary, source_dir, zonal_office_by_city):
	rows = _read_xlsx_rows(os.path.join(source_dir, "PM_Case_Outlets.xlsx"))
	for row in rows:
		code = normalize_whitespace(row.get("Outlet Code"))
		city = normalize_whitespace(row.get("City"))
		if not code or not city:
			summary.warn(f"CB Outlet: skipped row with missing code/city: {row}")
			continue
		if city not in OUTLET_CITIES:
			summary.warn(f"CB Outlet {code}: unrecognized city {city!r}, skipped")
			continue
		if frappe.db.exists("CB Outlet", code):
			summary.warn(f"CB Outlet {code}: already exists, skipped (collision check)")
			continue
		frappe.get_doc(
			{
				"doctype": "CB Outlet",
				"outlet_code": code,
				"city": city,
				"zonal_office": zonal_office_by_city[city],
				"status": "Active",
			}
		).insert()
		summary.count("CB Outlet created")


def import_asset_types(summary):
	canonical_types = sorted(set(ASSET_TYPE_ALIASES.values()))
	for name in canonical_types:
		if frappe.db.exists("CB Asset Type", name):
			continue
		frappe.get_doc({"doctype": "CB Asset Type", "asset_type_name": name, "active": 1}).insert()
		summary.count("CB Asset Type created")
	summary.count("CB Asset Type total canonical", len(canonical_types))
	return canonical_types


def import_technicians(summary, source_dir, zonal_office_by_city):
	path = os.path.join(source_dir, "PM_Case_User_Master.csv")
	with open(path, newline="", encoding="utf-8-sig") as fh:
		rows = list(csv.DictReader(fh))

	name_to_employee_no = {}
	excluded_names = set()

	# Pass 1: create technicians (reports_to deferred to pass 2), skip anyone whose
	# Home doesn't map to a real zonal office — zonal_office is required, and there's
	# no legitimate value to guess (see PROGRESS.md Phase 5: 'COR' / Corporate Office).
	for row in rows:
		employee_no = normalize_whitespace(row.get("Employee No"))
		name = normalize_whitespace(row.get("Name"))
		home = normalize_whitespace(row.get("Home"))
		city = HOME_TO_CITY.get(home)

		if frappe.db.exists("CB Technician", employee_no):
			summary.warn(f"CB Technician {employee_no} ({name}): already exists, skipped (collision check)")
			name_to_employee_no[name] = employee_no
			continue

		if not city:
			summary.warn(
				f"CB Technician {employee_no} ({name}): unrecognized Home {home!r} — cannot "
				f"derive zonal_office (required field), excluded from import rather than guessed"
			)
			excluded_names.add(name)
			continue

		frappe.get_doc(
			{
				"doctype": "CB Technician",
				"employee_no": employee_no,
				"technician_name": name,
				"job_title": normalize_whitespace(row.get("Job title")),
				"department": normalize_whitespace(row.get("Department")) or "Maintenance",
				"email": normalize_whitespace(row.get("Email")),
				"mobile": normalize_whitespace(row.get("Mobile")),
				"zonal_office": zonal_office_by_city[city],
				"active": 1,
			}
		).insert()
		name_to_employee_no[name] = employee_no
		summary.count("CB Technician created")

	summary.count("CB Technician excluded (unresolvable Home)", len(excluded_names))

	# Pass 2: resolve reports_to by fuzzy match, now that every real name exists.
	candidate_names = [n for n in name_to_employee_no if n not in excluded_names]
	for row in rows:
		name = normalize_whitespace(row.get("Name"))
		if name in excluded_names:
			continue
		employee_no = name_to_employee_no[name]
		if frappe.db.get_value("CB Technician", employee_no, "reports_to"):
			continue  # already resolved (e.g. re-run)
		raw_reports_to = row.get("Reports to")
		match, method = match_reports_to(raw_reports_to, candidate_names)
		if match:
			frappe.db.set_value("CB Technician", employee_no, "reports_to", name_to_employee_no[match])
			summary.count("CB Technician reports_to resolved")
		else:
			summary.warn(
				f"CB Technician {employee_no} ({name}): reports_to {raw_reports_to!r} left "
				f"blank — {method}"
			)
			summary.count("CB Technician reports_to unresolved (logged, left blank)")


def import_assets_from_before_sample(summary, source_dir):
	"""Synthesize CB Asset rows. PM_Case_Before.xlsx is a PM tracker export, not an
	asset register (docs/DocType_Spec.md section 5's own note) — rather than
	fabricate a full production-scale register across all 133 outlets, this creates
	exactly one representative asset per (outlet, canonical asset type) pair that
	ACTUALLY co-occurs in Before.xlsx's own rows. This is directly evidenced by the
	source data, not invented, and stays small (see PROGRESS.md Phase 5)."""
	rows = _read_xlsx_rows(os.path.join(source_dir, "PM_Case_Before.xlsx"), sheet_name="PM Tracker 2026")
	pairs = set()
	for row in rows:
		outlet = normalize_whitespace(row.get("Outlet"))
		asset_type = canonicalize_asset_type(row.get("Asset"))
		if outlet and asset_type:
			pairs.add((outlet, asset_type))

	for outlet, asset_type in sorted(pairs):
		if not frappe.db.exists("CB Outlet", outlet):
			summary.warn(f"CB Asset synthesis: outlet {outlet!r} not found in CB Outlet, skipped")
			continue
		if frappe.db.exists("CB Asset", {"outlet": outlet, "asset_type": asset_type}):
			summary.warn(
				f"CB Asset synthesis: {outlet}/{asset_type} already has an asset, skipped (collision check)"
			)
			continue
		frappe.get_doc(
			{"doctype": "CB Asset", "asset_type": asset_type, "outlet": outlet, "status": "Active"}
		).insert()  # CB Asset.after_insert fires — matching existing PM Programs get scheduled immediately
		summary.count("CB Asset synthesized (from Before.xlsx sample outlets)")


def import_pm_programs(summary, source_dir):
	rows = _read_xlsx_rows(os.path.join(source_dir, "PM_Case_Before.xlsx"), sheet_name="PM Tracker 2026")
	groups = {}
	for row in rows:
		asset_type = canonicalize_asset_type(row.get("Asset"))
		task = normalize_whitespace(row.get("Task"))
		freq = normalize_frequency(row.get("Freq"))
		groups.setdefault((asset_type, task), []).append(freq)

	for (asset_type, task), freqs in groups.items():
		non_blank = {f for f in freqs if f}
		blanks = sum(1 for f in freqs if not f)

		if len(non_blank) > 1:
			summary.warn(
				f"CB PM Program: conflicting frequencies for ({asset_type!r}, {task!r}): "
				f"{sorted(non_blank)} — not created"
			)
			summary.count("CB PM Program frequency conflicts")
			continue

		if len(non_blank) == 0:
			summary.warn(
				f"CB PM Program: no frequency anywhere for ({asset_type!r}, {task!r}), "
				f"{blanks} row(s) excluded"
			)
			summary.count("CB PM Program unresolved rows (excluded)", blanks)
			continue

		frequency = next(iter(non_blank))
		program_key = f"{asset_type or ''}|{task}|{frequency}"
		if blanks:
			summary.count("CB PM Program blank-freq rows resolved via group", blanks)
		if frappe.db.exists("CB PM Program", {"program_key": program_key}):
			continue

		label = f"{asset_type} - {task} - {frequency}" if asset_type else f"{task} - {frequency}"
		frappe.get_doc(
			{
				"doctype": "CB PM Program",
				"program_name": label,
				"asset_type": asset_type,
				"task_description": task,
				"frequency": frequency,
				"active": 1,
			}
		).insert()
		summary.count("CB PM Program created")


def import_ticket_taxonomy(summary, source_dir):
	rows = _read_xlsx_rows(os.path.join(source_dir, "PM_Case_Ticket_Buckets.xlsx"))
	seen_keys = set()
	for row in rows:
		department = normalize_whitespace(row.get("Department"))
		category = normalize_whitespace(row.get("Category"))
		sub1 = normalize_whitespace(row.get("Sub Category 1"))
		sub2 = normalize_whitespace(row.get("Sub Category 2"))
		if not department or not category:
			# e.g. one real row: ('Snags', None, None, None) — both department and
			# category are required on CB Ticket Taxonomy; nothing sensible to guess.
			summary.warn(f"CB Ticket Taxonomy: skipped row missing department/category: {row}")
			summary.count("CB Ticket Taxonomy skipped (missing department/category)")
			continue
		key = "|".join([department, category, sub1, sub2])
		if key in seen_keys:
			summary.warn(f"CB Ticket Taxonomy: duplicate row skipped: {(department, category, sub1, sub2)}")
			summary.count("CB Ticket Taxonomy duplicates skipped")
			continue
		seen_keys.add(key)
		if frappe.db.exists("CB Ticket Taxonomy", {"taxonomy_key": key}):
			continue
		frappe.get_doc(
			{
				"doctype": "CB Ticket Taxonomy",
				"department": department,
				"category": category,
				"sub_category_1": sub1,
				"sub_category_2": sub2,
			}
		).insert()
		summary.count("CB Ticket Taxonomy created")


def import_spare_parts(summary, source_dir):
	rows = _read_xlsx_rows(os.path.join(source_dir, "PM_Case_Ticket_Buckets.xlsx"))
	for row in rows:
		if normalize_whitespace(row.get("Department")) != "Spare Parts":
			continue
		equipment_model = normalize_whitespace(row.get("Category"))
		parsed = parse_spare_part(row.get("Sub Category 1"))
		if not parsed:
			summary.warn(f"CB Spare Part: unparseable Sub Category 1 {row.get('Sub Category 1')!r}, skipped")
			summary.count("CB Spare Part unparseable")
			continue
		part_code, part_name = parsed
		if frappe.db.exists("CB Spare Part", part_code):
			summary.warn(f"CB Spare Part {part_code}: already exists, skipped (collision check)")
			continue
		frappe.get_doc(
			{
				"doctype": "CB Spare Part",
				"part_code": part_code,
				"part_name": part_name,
				"equipment_model": equipment_model,
				"active": 1,
			}
		).insert()
		summary.count("CB Spare Part created")


def seed_initial_schedules(summary):
	"""The 'initial setup' trigger from docs/DocType_Spec.md section 7: for every
	active PM Program, find its applicable targets fresh and ensure_schedule each —
	regardless of whether the program or its targets were just created in this run or
	already existed (e.g. the Phase 1/2 fixture program, or BLR001/BLR134's assets)."""
	program_names = frappe.get_all("CB PM Program", filters={"active": 1}, pluck="name")
	for program_name in program_names:
		program = frappe.get_doc("CB PM Program", program_name)
		for outlet, asset in find_applicable_targets(program):
			result = ensure_schedule(program_name, outlet, asset, today())
			if not isinstance(result, str):  # a string means it already existed
				summary.count("CB PM Schedule seeded (new)")


def run(source_dir="/workspace-project/data/source"):
	frappe.set_user("Administrator")
	summary = ImportSummary()

	zonal_office_by_city = import_zonal_offices(summary)
	import_outlets(summary, source_dir, zonal_office_by_city)
	import_asset_types(summary)
	import_technicians(summary, source_dir, zonal_office_by_city)
	import_assets_from_before_sample(summary, source_dir)
	import_pm_programs(summary, source_dir)
	import_ticket_taxonomy(summary, source_dir)
	import_spare_parts(summary, source_dir)
	seed_initial_schedules(summary)

	frappe.db.commit()
	summary.print_report()
	return summary.counts, summary.warnings
